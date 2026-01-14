import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import os
import io
import re
import hashlib
import json
from datetime import datetime, date, time
from pandas.errors import EmptyDataError

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

# Optional (only needed for Chatbot view)
try:
    from openai import OpenAI
except Exception:
    OpenAI = None

# =========================================================
# PERSISTENT DATABASE (STREAMLIT CLOUD SAFE)
# =========================================================
DB_DIR = "/mount/src/.data"
os.makedirs(DB_DIR, exist_ok=True)
DB_PATH = os.path.join(DB_DIR, "smt_spit.db")

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(page_title="SMT Spit Analytics", layout="wide")

REJECT_CODES = {2, 3, 4, 5, 6, 7}

# Board counting model
LINE1_MACHINES = {"EPS16"}
LINE2_MACHINES = {"IINEO682", "IIN2-053-2", "IIN2-053-1"}
LINE2_DIVISOR = 3  # line2 boards estimated as logs/3

# Master BOM format
MASTER_BOM_COMP_COL_INDEX = 0  # Column A
MASTER_BOM_COST_COL_INDEX = 9  # Column J

# Filename example: 20260106091251-IIN2-053-2.csv
FILENAME_RE = re.compile(r"^(?P<dt>\d{14})-(?P<machine>.+?)(?:\.[A-Za-z0-9]+)?$")

# =========================================================
# DB HELPERS
# =========================================================
def db_connect():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn

def _table_has_column(conn: sqlite3.Connection, table: str, col: str) -> bool:
    cols = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(c[1] == col for c in cols)

def db_init(conn: sqlite3.Connection):
    conn.execute("""
    CREATE TABLE IF NOT EXISTS bom_versions (
        bom_id INTEGER PRIMARY KEY AUTOINCREMENT,
        bom_name TEXT NOT NULL,
        uploaded_at TEXT NOT NULL
    )
    """)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS bom_items (
        bom_id INTEGER NOT NULL,
        component TEXT NOT NULL,
        unit_cost REAL NOT NULL,
        PRIMARY KEY (bom_id, component),
        FOREIGN KEY(bom_id) REFERENCES bom_versions(bom_id) ON DELETE CASCADE
    )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_bom_items_component ON bom_items(component)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_bom_items_bomid ON bom_items(bom_id)")

    conn.execute("""
    CREATE TABLE IF NOT EXISTS logs (
        file_hash TEXT PRIMARY KEY,
        filename TEXT NOT NULL,
        file_dt TEXT,
        machine TEXT,
        board_name TEXT,
        mo TEXT,
        ingested_at TEXT NOT NULL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_hash TEXT NOT NULL,
        component TEXT,
        description TEXT,
        location TEXT,
        board_name TEXT,
        mo TEXT,
        file_dt TEXT,
        machine TEXT,
        unit_cost REAL,
        cost REAL
    )
    """)

    # ---- MIGRATION: add reject_code column if missing
    if not _table_has_column(conn, "events", "reject_code"):
        conn.execute("ALTER TABLE events ADD COLUMN reject_code INTEGER")

    # ---- MIGRATION: add feeder_no / slot_no columns if missing (H, I)
    if not _table_has_column(conn, "events", "feeder_no"):
        conn.execute("ALTER TABLE events ADD COLUMN feeder_no TEXT")
    if not _table_has_column(conn, "events", "slot_no"):
        conn.execute("ALTER TABLE events ADD COLUMN slot_no TEXT")

    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_dt ON events(file_dt)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_board ON events(board_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_mo ON events(mo)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_machine ON events(machine)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_comp ON events(component)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_reject_code ON events(reject_code)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_feeder ON events(feeder_no)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_events_slot ON events(slot_no)")

    conn.commit()

def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

# =========================================================
# PARSERS
# =========================================================
def parse_cost(v):
    if pd.isna(v):
        return np.nan
    if isinstance(v, (int, float, np.number)):
        return float(v)
    s = str(v).replace(",", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except:
        return np.nan

def parse_dt_machine_from_filename(filename: str):
    base = os.path.basename(filename)
    m = FILENAME_RE.match(base)
    if not m:
        return None, None
    dt_str = m.group("dt")
    machine = m.group("machine")
    try:
        dt = datetime.strptime(dt_str, "%Y%m%d%H%M%S")
        return dt.isoformat(sep=" "), machine
    except:
        return None, machine

def safe_read_csv(bytes_data, **kwargs):
    try:
        return pd.read_csv(io.BytesIO(bytes_data), encoding="utf-8", **kwargs)
    except UnicodeDecodeError:
        return pd.read_csv(io.BytesIO(bytes_data), encoding="latin-1", **kwargs)
    except EmptyDataError:
        raise

def _clean_cell(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none", "null"}:
        return None
    return s

def read_header_board_mo(file_bytes: bytes, filename: str):
    ext = filename.lower().split(".")[-1]
    try:
        if ext in ("xls", "xlsx"):
            header = pd.read_excel(io.BytesIO(file_bytes), nrows=1, header=None)
        else:
            header = safe_read_csv(file_bytes, nrows=1, header=None)
    except Exception:
        return None, None

    try:
        board = _clean_cell(header.iloc[0, 1])  # B1
        mo = _clean_cell(header.iloc[0, 3])     # D1
        return board, mo
    except Exception:
        return None, None

def read_body_df(file_bytes: bytes, filename: str):
    ext = filename.lower().split(".")[-1]
    try:
        if ext in ("xls", "xlsx"):
            df = pd.read_excel(io.BytesIO(file_bytes), skiprows=2, header=None, usecols=range(12))
        else:
            df = safe_read_csv(file_bytes, skiprows=2, header=None, usecols=range(12))
    except EmptyDataError:
        return None
    except Exception:
        return None

    if df is None or df.empty:
        return None
    df = df.iloc[:, :12]
    if df.shape[1] < 12:
        return None
    df.columns = list("ABCDEFGHIJKL")
    return df

# =========================================================
# BOM (VERSIONED, MASTER BOM A/J across all sheets)
# =========================================================
def ingest_master_bom(conn: sqlite3.Connection, bom_bytes: bytes, bom_name: str) -> int:
    xls = pd.ExcelFile(io.BytesIO(bom_bytes))
    items = []

    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        if df.shape[1] <= max(MASTER_BOM_COMP_COL_INDEX, MASTER_BOM_COST_COL_INDEX):
            continue

        for i in range(len(df)):
            comp = _clean_cell(df.iat[i, MASTER_BOM_COMP_COL_INDEX])
            if not comp:
                continue
            if comp.lower() in {"component", "part", "part number", "item", "sku"}:
                continue

            cost = parse_cost(df.iat[i, MASTER_BOM_COST_COL_INDEX])
            if pd.isna(cost):
                continue

            items.append((comp, float(cost)))

    if not items:
        return 0

    uploaded_at = datetime.now().isoformat(sep=" ")
    cur = conn.execute(
        "INSERT INTO bom_versions(bom_name, uploaded_at) VALUES (?, ?)",
        (bom_name, uploaded_at)
    )
    bom_id = cur.lastrowid

    tmp = {}
    for comp, cost in items:
        tmp[comp] = cost

    conn.executemany(
        "INSERT INTO bom_items(bom_id, component, unit_cost) VALUES (?,?,?)",
        [(bom_id, c, tmp[c]) for c in tmp.keys()]
    )
    conn.commit()
    return len(tmp)

def list_boms(conn: sqlite3.Connection) -> pd.DataFrame:
    return pd.read_sql_query(
        "SELECT bom_id, bom_name, uploaded_at FROM bom_versions ORDER BY bom_id DESC",
        conn
    )

def get_bom_lookup(conn: sqlite3.Connection, selected_bom_ids: list[int] | None) -> dict:
    if not selected_bom_ids:
        sql = """
        SELECT bi.component, bi.unit_cost
        FROM bom_items bi
        JOIN (
            SELECT component, MAX(bom_id) AS max_bom_id
            FROM bom_items
            GROUP BY component
        ) latest
        ON bi.component = latest.component AND bi.bom_id = latest.max_bom_id
        """
        rows = conn.execute(sql).fetchall()
        return {r[0]: float(r[1]) for r in rows}

    placeholders = ",".join(["?"] * len(selected_bom_ids))
    sql = f"""
    SELECT bi.component, bi.unit_cost
    FROM bom_items bi
    JOIN (
        SELECT component, MAX(bom_id) AS max_bom_id
        FROM bom_items
        WHERE bom_id IN ({placeholders})
        GROUP BY component
    ) latest
    ON bi.component = latest.component AND bi.bom_id = latest.max_bom_id
    """
    rows = conn.execute(sql, selected_bom_ids).fetchall()
    return {r[0]: float(r[1]) for r in rows}

# =========================================================
# INGEST LOGS
# =========================================================
def ingest_logs(conn: sqlite3.Connection, uploads):
    # Always use latest-per-component for ingestion; analysis can use selected BOM versions
    bom_lookup = get_bom_lookup(conn, selected_bom_ids=None)

    skipped = []
    inserted_files = 0
    inserted_events = 0

    for up in uploads:
        filename = up.name
        b = up.getvalue()

        if not b:
            skipped.append((filename, "Empty file"))
            continue

        file_hash = sha256_bytes(b)
        if conn.execute("SELECT 1 FROM logs WHERE file_hash=?", (file_hash,)).fetchone():
            skipped.append((filename, "Already ingested (same hash)"))
            continue

        dt_iso, machine = parse_dt_machine_from_filename(filename)
        board, mo = read_header_board_mo(b, filename)

        df = read_body_df(b, filename)
        if df is None:
            skipped.append((filename, "No readable data after skiprows=2"))
            continue

        conn.execute(
            "INSERT INTO logs(file_hash, filename, file_dt, machine, board_name, mo, ingested_at) VALUES (?,?,?,?,?,?,?)",
            (file_hash, filename, dt_iso, machine, board, mo, datetime.now().isoformat(sep=" "))
        )
        inserted_files += 1

        ev_rows = []
        for _, r in df.iterrows():
            try:
                code = int(r["L"])
            except Exception:
                continue

            if code in REJECT_CODES:
                try:
                    comp = str(r["B"]).strip()
                    desc = str(r["C"]).strip()
                    loc = str(r["D"]).strip()
                    feeder = str(r["H"]).strip()  # Column H
                    slot = str(r["I"]).strip()    # Column I
                    cost = float(bom_lookup.get(comp, 0.0))

                    ev_rows.append((
                        file_hash, comp, desc, loc, feeder, slot, board, mo, dt_iso, machine, cost, cost, code
                    ))
                except Exception:
                    continue

        if ev_rows:
            conn.executemany(
                """
                INSERT INTO events(
                    file_hash, component, description, location, feeder_no, slot_no,
                    board_name, mo, file_dt, machine, unit_cost, cost, reject_code
                )
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                ev_rows
            )
            inserted_events += len(ev_rows)

        conn.commit()

    return inserted_files, inserted_events, skipped

# =========================================================
# FILTER SQL
# =========================================================
def _build_where(dt_start, dt_end, boards, mos, machines, components=None):
    where = []
    params = []

    if dt_start is not None:
        where.append("(file_dt IS NOT NULL AND file_dt >= ?)")
        params.append(dt_start.isoformat(sep=" "))
    if dt_end is not None:
        where.append("(file_dt IS NOT NULL AND file_dt <= ?)")
        params.append(dt_end.isoformat(sep=" "))

    if boards:
        where.append("board_name IN (%s)" % ",".join(["?"] * len(boards)))
        params.extend(boards)
    if mos:
        where.append("mo IN (%s)" % ",".join(["?"] * len(mos)))
        params.extend(mos)
    if machines:
        where.append("machine IN (%s)" % ",".join(["?"] * len(machines)))
        params.extend(machines)
    if components is not None and components:
        where.append("component IN (%s)" % ",".join(["?"] * len(components)))
        params.extend(components)

    return where, params

# =========================================================
# BOARD COUNT (LINE2 = logs/3)
# =========================================================
def estimate_total_boards(conn, dt_start, dt_end, boards, mos, machines) -> float:
    where, params = _build_where(dt_start, dt_end, boards, mos, machines, components=None)
    sql = "SELECT machine, COUNT(*) AS n FROM logs"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY machine"
    df = pd.read_sql_query(sql, conn, params=params)

    if df.empty:
        return 0.0

    line1_logs = float(df.loc[df["machine"].isin(LINE1_MACHINES), "n"].sum())
    line2_logs = float(df.loc[df["machine"].isin(LINE2_MACHINES), "n"].sum())
    other_logs = float(df.loc[~df["machine"].isin(LINE1_MACHINES.union(LINE2_MACHINES)), "n"].sum())

    return line1_logs + (line2_logs / LINE2_DIVISOR) + other_logs

def estimate_boards_by_board(conn, dt_start, dt_end, boards, mos, machines, boards_limit=None) -> pd.DataFrame:
    where, params = _build_where(dt_start, dt_end, boards, mos, machines, components=None)
    sql = "SELECT board_name AS Board, machine AS Machine, COUNT(*) AS n FROM logs"
    if where:
        sql += " WHERE " + " AND ".join(where)
    if boards_limit:
        if where:
            sql += " AND "
        else:
            sql += " WHERE "
        sql += "board_name IN (%s)" % ",".join(["?"] * len(boards_limit))
        params.extend(boards_limit)
    sql += " GROUP BY board_name, machine"
    df = pd.read_sql_query(sql, conn, params=params)

    if df.empty:
        return pd.DataFrame(columns=["Board", "BoardsRun"])

    def weight(m):
        if m in LINE1_MACHINES:
            return 1.0
        if m in LINE2_MACHINES:
            return 1.0 / LINE2_DIVISOR
        return 1.0

    df["BoardsEquivalent"] = df.apply(lambda r: float(r["n"]) * weight(r["Machine"]), axis=1)
    out = df.groupby("Board")["BoardsEquivalent"].sum().reset_index().rename(columns={"BoardsEquivalent": "BoardsRun"})
    return out

def machine_log_breakdown(conn, dt_start, dt_end, boards, mos, machines) -> pd.DataFrame:
    where, params = _build_where(dt_start, dt_end, boards, mos, machines, components=None)
    sql = "SELECT machine AS Machine, COUNT(*) AS LogFiles FROM logs"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY machine ORDER BY LogFiles DESC"
    return pd.read_sql_query(sql, conn, params=params)

# =========================================================
# EVENTS QUERY (includes Reject Code + Feeder/Slot)
# =========================================================
def query_events(conn, dt_start, dt_end, boards, mos, machines, components, bom_lookup):
    where, params = _build_where(dt_start, dt_end, boards, mos, machines, components)
    sql = """
    SELECT
      component AS Component,
      description AS Description,
      location AS Location,
      feeder_no AS Feeder,
      slot_no AS Slot,
      board_name AS Board,
      mo AS MO,
      file_dt AS FileDateTime,
      machine AS Machine,
      reject_code AS RejectCode
    FROM events
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY file_dt DESC"
    df = pd.read_sql_query(sql, conn, params=params)

    if df.empty:
        df["UnitCost"] = []
        df["Cost"] = []
        return df

    df["UnitCost"] = df["Component"].map(lambda c: float(bom_lookup.get(str(c).strip(), 0.0)))
    df["Cost"] = df["UnitCost"]
    return df

# =========================================================
# DERIVED VIEWS (Summary includes Reject Codes breakdown)
# =========================================================
def _format_reject_codes(series: pd.Series) -> str:
    s = series.dropna()
    if s.empty:
        return ""
    try:
        s = s.astype(int)
    except Exception:
        s = pd.to_numeric(s, errors="coerce").dropna().astype(int)
        if s.empty:
            return ""
    vc = s.value_counts()
    parts = []
    for code, cnt in vc.sort_values(ascending=False).items():
        parts.append(f"{int(cnt)}x C{int(code)}")
    return ", ".join(parts)

def make_summary(events_df):
    if events_df.empty:
        return pd.DataFrame(columns=["Component","Description","Machine","Spits","Reject Codes","UnitCost","TotalCost"])
    return (
        events_df.groupby("Component")
        .agg(
            Description=("Description", lambda x: x.mode().iloc[0] if len(x.mode()) else x.iloc[0]),
            Machine=("Machine", lambda x: x.mode().iloc[0] if len(x.mode()) else x.iloc[0]),
            Spits=("Component", "count"),
            RejectCodes=("RejectCode", _format_reject_codes),
            UnitCost=("UnitCost", "max"),
            TotalCost=("Cost", "sum"),
        )
        .reset_index()
        .rename(columns={"RejectCodes": "Reject Codes"})
        .sort_values("TotalCost", ascending=False)
    )

def make_repeated_locations(events_df):
    if events_df.empty:
        return pd.DataFrame(columns=["Component","Location","Board","Machine","Spits","TotalCost"])
    return (
        events_df.groupby(["Component","Location","Board","Machine"])
        .agg(Spits=("Component","count"), TotalCost=("Cost","sum"))
        .reset_index()
        .query("Spits > 1")
        .sort_values(["TotalCost","Spits"], ascending=[False, False])
    )

def make_missing_costs(events_df):
    if events_df.empty:
        return pd.DataFrame(columns=["Component","Spits (cost=0)"])
    return (
        events_df.loc[events_df["UnitCost"] == 0.0, "Component"]
        .value_counts()
        .reset_index()
        .rename(columns={"index":"Component", "Component":"Spits (cost=0)"})
    )

def make_board_loss(events_df, board_value):
    if events_df.empty:
        return pd.DataFrame(columns=["Board","TotalCost","Loss % of Board Value (period)"])
    out = events_df.groupby("Board")["Cost"].sum().reset_index(name="TotalCost")
    out["Loss % of Board Value (period)"] = (out["TotalCost"] / board_value * 100) if board_value else np.nan
    return out.sort_values("Loss % of Board Value (period)", ascending=False)

def make_board_loss_components(events_df, boards_run_by_board, board_value):
    if events_df.empty:
        return pd.DataFrame(columns=[
            "Board","BoardsRun","Component","Description","Spits","UnitCost","Cost",
            "Period % of Board Value","Avg % of Board Value per Board","% of Board’s Total Loss"
        ])

    comp_loss = (
        events_df.groupby(["Board","Component"])
        .agg(
            Description=("Description", lambda x: x.mode().iloc[0] if len(x.mode()) else x.iloc[0]),
            Spits=("Component","count"),
            UnitCost=("UnitCost","max"),
            Cost=("Cost","sum")
        )
        .reset_index()
    )

    comp_loss = comp_loss.merge(boards_run_by_board, on="Board", how="left")
    comp_loss["BoardsRun"] = comp_loss["BoardsRun"].fillna(0.0)

    comp_loss["Period % of Board Value"] = ((comp_loss["Cost"] / board_value) * 100) if board_value else np.nan
    comp_loss["Avg % of Board Value per Board"] = np.where(
        (board_value > 0) & (comp_loss["BoardsRun"] > 0),
        (comp_loss["Cost"] / (board_value * comp_loss["BoardsRun"])) * 100,
        np.nan
    )

    board_total = comp_loss.groupby("Board")["Cost"].transform("sum")
    comp_loss["% of Board’s Total Loss"] = np.where(board_total > 0, (comp_loss["Cost"] / board_total) * 100, 0.0)

    return comp_loss.sort_values(["Board", "Cost"], ascending=[True, False])

# =========================================================
# EXCEL EXPORT
# =========================================================
def _fit_columns(ws, max_width=60):
    for col_cells in ws.columns:
        length = 0
        col_letter = col_cells[0].column_letter
        for c in col_cells:
            if c.value is None:
                continue
            length = max(length, len(str(c.value)))
        ws.column_dimensions[col_letter].width = min(max(10, length + 2), max_width)

def _add_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

def build_excel_report(
    events_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    pareto_df: pd.DataFrame,
    repeated_df: pd.DataFrame,
    yield_df: pd.DataFrame,
    missing_df: pd.DataFrame,
    board_loss_df: pd.DataFrame,
    board_loss_components_df: pd.DataFrame
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Summary
    ws = wb.create_sheet("Summary")
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    if ws.max_row >= 2:
        _add_table(ws, "SummaryTable", f"A1:G{ws.max_row}")
    _fit_columns(ws)

    # ---- Spit Events (now includes Feeder + Slot)
    ws = wb.create_sheet("Spit Events")
    for r in dataframe_to_rows(events_df, index=False, header=True):
        ws.append(r)
    if ws.max_row >= 2:
        # events_df columns: Component, Description, Location, Feeder, Slot, Board, MO, FileDateTime, Machine, RejectCode, UnitCost, Cost = 12 columns
        _add_table(ws, "SpitEventsTable", f"A1:L{ws.max_row}")
    _fit_columns(ws)

    # ---- Pareto
    ws = wb.create_sheet("Pareto (Cost)")
    ws.append(["Component", "TotalCost"])
    for _, row in pareto_df.iterrows():
        ws.append([row["Component"], float(row["TotalCost"])])
    _fit_columns(ws)

    if len(pareto_df) > 0:
        chart = BarChart()
        chart.title = "Cost-Based Pareto (Top Offenders)"
        chart.y_axis.title = "Total Cost Loss"
        chart.x_axis.title = "Component"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(pareto_df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(pareto_df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 28
        ws.add_chart(chart, "D2")

    # ---- Repeated Locations
    ws = wb.create_sheet("Repeated Locations")
    for r in dataframe_to_rows(repeated_df, index=False, header=True):
        ws.append(r)
    if ws.max_row >= 2 and repeated_df.shape[1] > 0:
        _add_table(ws, "RepeatedLocationsTable", f"A1:F{ws.max_row}")
    _fit_columns(ws)

    # ---- Yield Loss
    ws = wb.create_sheet("Yield Loss")
    for r in dataframe_to_rows(yield_df, index=False, header=True):
        ws.append(r)
    _fit_columns(ws)

    # ---- Missing BOM Costs
    ws = wb.create_sheet("Missing BOM Costs")
    for r in dataframe_to_rows(missing_df, index=False, header=True):
        ws.append(r)
    if ws.max_row >= 2 and missing_df.shape[1] > 0:
        _add_table(ws, "MissingCostTable", f"A1:B{ws.max_row}")
    _fit_columns(ws)

    # ---- Board Loss %
    ws = wb.create_sheet("Board Loss %")
    for r in dataframe_to_rows(board_loss_df, index=False, header=True):
        ws.append(r)
    if ws.max_row >= 2 and board_loss_df.shape[1] > 0:
        _add_table(ws, "BoardLossTable", f"A1:C{ws.max_row}")
    _fit_columns(ws)

    # ---- Board Loss Components
    ws = wb.create_sheet("Board Loss Components")
    for r in dataframe_to_rows(board_loss_components_df, index=False, header=True):
        ws.append(r)
    if ws.max_row >= 2 and board_loss_components_df.shape[1] > 0:
        _add_table(ws, "BoardLossComponentsTable", f"A1:J{ws.max_row}")
    _fit_columns(ws)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# =========================================================
# RESET DB (ADMIN)
# =========================================================
def reset_database():
    try:
        if os.path.exists(DB_PATH):
            os.remove(DB_PATH)
    except Exception as e:
        return False, str(e)
    return True, None

# =========================================================
# CHATBOT (OPTION 2: Grounded + Advisor)
# =========================================================
CHAT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "last_run",
            "description": "Get the most recent datetime a board was run (based on ingested logs).",
            "parameters": {
                "type": "object",
                "properties": {
                    "board": {"type": "string"}
                },
                "required": ["board"]
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "boards_run",
            "description": "Estimate boards run in a datetime window, using line-2 ÷3 correction. Optional filters: board, mo, machine.",
            "parameters": {
                "type": "object",
                "properties": {
                    "start": {"type": "string", "description": "ISO datetime 'YYYY-MM-DD HH:MM:SS'"},
                    "end": {"type": "string", "description": "ISO datetime 'YYYY-MM-DD HH:MM:SS'"},
                    "board": {"type": ["string", "null"]},
                    "mo": {"type": ["string", "null"]},
                    "machine": {"type": ["string", "null"]},
                },
                "required": ["start", "end"]
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "top_offenders",
            "description": "Top offending components by spit count or cost within a datetime window. Optional filters: board, mo, machine.",
            "parameters": {
                "type": "object",
                "properties": {
                    "start": {"type": "string"},
                    "end": {"type": "string"},
                    "by": {"type": "string", "enum": ["count", "cost"]},
                    "board": {"type": ["string", "null"]},
                    "mo": {"type": ["string", "null"]},
                    "machine": {"type": ["string", "null"]},
                    "limit": {"type": "integer", "minimum": 1, "maximum": 50, "default": 10},
                },
                "required": ["start", "end", "by"]
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "worst_feeder_slot",
            "description": "Worst feeder/slot combinations by count or cost within a datetime window. Optional filter: machine.",
            "parameters": {
                "type": "object",
                "properties": {
                    "start": {"type": "string"},
                    "end": {"type": "string"},
                    "by": {"type": "string", "enum": ["count", "cost"]},
                    "machine": {"type": ["string", "null"]},
                    "limit": {"type": "integer", "minimum": 1, "maximum": 50, "default": 15},
                },
                "required": ["start", "end", "by"]
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "reject_code_breakdown",
            "description": "Reject code breakdown for a component within a datetime window. Optional filters: board, mo, machine.",
            "parameters": {
                "type": "object",
                "properties": {
                    "start": {"type": "string"},
                    "end": {"type": "string"},
                    "component": {"type": "string"},
                    "board": {"type": ["string", "null"]},
                    "mo": {"type": ["string", "null"]},
                    "machine": {"type": ["string", "null"]},
                },
                "required": ["start", "end", "component"]
            },
        },
    },
]

def tool_last_run(conn, board: str):
    df = pd.read_sql_query(
        "SELECT MAX(file_dt) AS last_dt FROM logs WHERE board_name=?",
        conn, params=[board]
    )
    return {"board": board, "last_dt": df.loc[0, "last_dt"]}

def tool_boards_run(conn, start: str, end: str, board=None, mo=None, machine=None):
    dt_start = datetime.fromisoformat(start)
    dt_end = datetime.fromisoformat(end)
    boards_sel = [board] if board else []
    mos_sel = [mo] if mo else []
    machines_sel = [machine] if machine else []
    n = estimate_total_boards(conn, dt_start, dt_end, boards_sel, mos_sel, machines_sel)
    return {
        "start": start, "end": end,
        "board": board, "mo": mo, "machine": machine,
        "estimated_boards": float(n),
        "line2_divisor": LINE2_DIVISOR
    }

def tool_top_offenders(conn, start: str, end: str, by: str, board=None, mo=None, machine=None, limit=10, selected_bom_ids=None):
    dt_start = datetime.fromisoformat(start)
    dt_end = datetime.fromisoformat(end)
    boards_sel = [board] if board else []
    mos_sel = [mo] if mo else []
    machines_sel = [machine] if machine else []

    bom_lookup = get_bom_lookup(conn, selected_bom_ids if selected_bom_ids else None)
    df = query_events(conn, dt_start, dt_end, boards_sel, mos_sel, machines_sel, components=[], bom_lookup=bom_lookup)
    if df.empty:
        return {"rows": []}

    limit = int(limit) if limit else 10
    if by == "count":
        out = (df.groupby("Component")
               .size().reset_index(name="Spits")
               .sort_values("Spits", ascending=False)
               .head(limit))
        return {"rows": out.to_dict(orient="records")}
    else:
        out = (df.groupby("Component")["Cost"]
               .sum().reset_index(name="TotalCost")
               .sort_values("TotalCost", ascending=False)
               .head(limit))
        return {"rows": out.to_dict(orient="records")}

def tool_worst_feeder_slot(conn, start: str, end: str, by: str, machine=None, limit=15):
    where = ["file_dt >= ?", "file_dt <= ?"]
    params = [start, end]
    if machine:
        where.append("machine = ?")
        params.append(machine)
    where_sql = "WHERE " + " AND ".join(where)
    limit = int(limit) if limit else 15

    if by == "count":
        df = pd.read_sql_query(
            f"""
            SELECT machine AS Machine, feeder_no AS Feeder, slot_no AS Slot,
                   component AS Component, COUNT(*) AS Spits
            FROM events
            {where_sql}
            GROUP BY machine, feeder_no, slot_no, component
            ORDER BY Spits DESC
            LIMIT {limit}
            """,
            conn, params=params
        )
        return {"rows": df.to_dict(orient="records")}
    else:
        df = pd.read_sql_query(
            f"""
            SELECT machine AS Machine, feeder_no AS Feeder, slot_no AS Slot,
                   component AS Component, SUM(cost) AS TotalCost
            FROM events
            {where_sql}
            GROUP BY machine, feeder_no, slot_no, component
            ORDER BY TotalCost DESC
            LIMIT {limit}
            """,
            conn, params=params
        )
        return {"rows": df.to_dict(orient="records")}

def tool_reject_code_breakdown(conn, start: str, end: str, component: str, board=None, mo=None, machine=None):
    where = ["file_dt >= ?", "file_dt <= ?", "component = ?"]
    params = [start, end, component]
    if board:
        where.append("board_name = ?"); params.append(board)
    if mo:
        where.append("mo = ?"); params.append(mo)
    if machine:
        where.append("machine = ?"); params.append(machine)
    where_sql = "WHERE " + " AND ".join(where)

    df = pd.read_sql_query(
        f"""
        SELECT reject_code AS Code, COUNT(*) AS Count
        FROM events
        {where_sql}
        GROUP BY reject_code
        ORDER BY Count DESC
        """,
        conn, params=params
    )
    return {"rows": df.to_dict(orient="records")}

def run_tool_by_name(conn, name: str, args: dict, selected_bom_ids=None):
    if name == "last_run":
        return tool_last_run(conn, **args)
    if name == "boards_run":
        return tool_boards_run(conn, **args)
    if name == "top_offenders":
        return tool_top_offenders(conn, selected_bom_ids=selected_bom_ids, **args)
    if name == "worst_feeder_slot":
        return tool_worst_feeder_slot(conn, **args)
    if name == "reject_code_breakdown":
        return tool_reject_code_breakdown(conn, **args)
    return {"error": f"Unknown tool: {name}"}

def chatbot_reply(conn, messages, default_start_iso: str, default_end_iso: str, advisor_mode: bool, selected_bom_ids=None):
    if OpenAI is None:
        return {"role": "assistant", "content": "The `openai` package is not installed. Add `openai` to requirements.txt and redeploy."}

    api_key = st.secrets.get("OPENAI_API_KEY", None)
    if not api_key:
        return {"role": "assistant", "content": "OPENAI_API_KEY is not set in Streamlit Secrets. Add it there to enable the chatbot."}

    client = OpenAI(api_key=api_key)

    system = (
        "You are a manufacturing analytics assistant for SMT pick-and-place log analysis.\n"
        "CRITICAL RULES:\n"
        "- For factual numbers/tables, always use tool outputs. Never invent values.\n"
        "- If the user doesn't specify a date range, use the provided defaults.\n"
        "- Write answers in two sections:\n"
        "  1) Facts (from data) — cite time range and filters used.\n"
        "  2) Suggestions (engineering judgement) — only if Advisor Mode is ON.\n"
        "- Keep it concise and actionable.\n"
    )

    defaults = (
        f"Default start: {default_start_iso}\n"
        f"Default end: {default_end_iso}\n"
        f"Advisor Mode: {advisor_mode}\n"
        f"Line2 board estimation divisor: {LINE2_DIVISOR}\n"
        "Note: Reject codes are C2..C7; feeder is column H; slot is column I.\n"
    )

    input_msgs = [{"role": "system", "content": system},
                  {"role": "system", "content": defaults}]
    input_msgs.extend(messages)

    # Step 1: model decides tool calls
    resp = client.responses.create(
        model="gpt-4.1",
        input=input_msgs,
        tools=CHAT_TOOLS,
        tool_choice="auto",
    )

    tool_calls = [o for o in resp.output if getattr(o, "type", None) == "function_call"]

    # If no tool calls, return any text it produced (usually a clarification question)
    if not tool_calls:
        text_parts = []
        for o in resp.output:
            if getattr(o, "type", None) == "output_text":
                text_parts.append(o.text)
        text = "\n".join(text_parts).strip()
        if not text:
            text = "I couldn't determine which data query to run. Try asking about boards run, last run, top offenders, feeder/slot, or reject codes."
        return {"role": "assistant", "content": text}

    # Execute tools
    tool_outputs = []
    debug_calls = []
    for call in tool_calls:
        name = call.name
        args = call.arguments if isinstance(call.arguments, dict) else json.loads(call.arguments)

        # Fill defaults if missing
        if "start" in args and not args["start"]:
            args["start"] = default_start_iso
        if "end" in args and not args["end"]:
            args["end"] = default_end_iso
        if "start" not in args and name != "last_run":
            args["start"] = default_start_iso
        if "end" not in args and name != "last_run":
            args["end"] = default_end_iso

        result = run_tool_by_name(conn, name, args, selected_bom_ids=selected_bom_ids)
        debug_calls.append({"tool": name, "args": args, "result_preview": result if isinstance(result, dict) else {"result": str(result)[:500]}})

        tool_outputs.append({
            "type": "function_call_output",
            "call_id": call.call_id,
            "output": json.dumps(result, default=str),
        })

    # Step 2: model writes final answer using tool outputs
    resp2 = client.responses.create(
        model="gpt-4.1",
        input=input_msgs,
        tools=CHAT_TOOLS,
        tool_choice="none",
        previous_response_id=resp.id,
        tool_outputs=tool_outputs,
    )

    text_parts = []
    for o in resp2.output:
        if getattr(o, "type", None) == "output_text":
            text_parts.append(o.text)
    text = "\n".join(text_parts).strip()
    if not text:
        text = "I ran the data queries but didn't get a textual response. Try again with a simpler question."

    return {"role": "assistant", "content": text, "debug_calls": debug_calls}

# =========================================================
# APP UI
# =========================================================
conn = db_connect()
db_init(conn)

st.title("SMT Spit Analytics (Full, Stable)")

# ---- Admin tools (Reset DB)
with st.expander("🔐 Admin Tools (Reset Database)"):
    admin_pw = st.text_input("Admin password", type="password")
    secret_pw = st.secrets.get("ADMIN_PASSWORD", None)

    if secret_pw is None:
        st.warning("Reset disabled: ADMIN_PASSWORD is not set in Streamlit Secrets.")
    else:
        st.caption("This deletes the saved database (BOMs + logs + events). Use carefully.")
        confirm = st.checkbox("I understand this will DELETE all saved data.")
        if st.button("RESET DATABASE", type="secondary"):
            if not confirm:
                st.error("Please tick the confirmation checkbox first.")
            elif admin_pw != secret_pw:
                st.error("Incorrect admin password.")
            else:
                ok, err = reset_database()
                if ok:
                    st.success("Database deleted. The app will rebuild tables on next run.")
                    st.session_state.clear()
                    st.rerun()
                else:
                    st.error(f"Reset failed: {err}")

# ---- View uploaded files and BOMs
with st.expander("📚 View uploaded Logs and Master BOM versions"):
    cA, cB = st.columns(2)
    with cA:
        st.subheader("BOM versions uploaded")
        boms_df = list_boms(conn)
        if boms_df.empty:
            st.info("No BOM versions stored yet.")
        else:
            st.dataframe(boms_df, use_container_width=True, height=240)

    with cB:
        st.subheader("Log files ingested (latest 200)")
        logs_df = pd.read_sql_query(
            "SELECT filename, file_dt, machine, board_name, mo, ingested_at FROM logs ORDER BY ingested_at DESC LIMIT 200",
            conn
        )
        if logs_df.empty:
            st.info("No logs ingested yet.")
        else:
            st.dataframe(logs_df, use_container_width=True, height=240)

# ---- Upload/Ingest
with st.expander("📦 Data Store (upload once, reused later)", expanded=True):
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Upload Master BOM (stored as new version)")
        bom_up = st.file_uploader(
            "Master BOM Excel: component in column A, cost in column J (all sheets read).",
            type=["xls", "xlsx"],
            key="bom_up"
        )
        bom_name = st.text_input("BOM name/label (e.g. Master BOM Jan-2026)", value="", key="bom_name")

        if st.button("Save Master BOM to Database", type="secondary"):
            if not bom_up:
                st.warning("Upload a Master BOM file first.")
            else:
                label = bom_name.strip() if bom_name.strip() else bom_up.name
                n = ingest_master_bom(conn, bom_up.getvalue(), label)
                if n == 0:
                    st.error("No BOM items were loaded. Check: component in column A and cost in column J.")
                else:
                    st.success(f"Master BOM stored. Components loaded: {n}")

    with col2:
        st.subheader("Ingest Log Files")
        logs_up = st.file_uploader(
            "Upload SMT log files (CSV/XLS/XLSX).",
            type=["csv", "xls", "xlsx"],
            accept_multiple_files=True,
            key="logs_up"
        )
        if st.button("Ingest Logs into Database", type="secondary"):
            if not logs_up:
                st.warning("Upload log files first.")
            else:
                ins_files, ins_events, skipped = ingest_logs(conn, logs_up)
                st.success(f"Ingest complete. New files: {ins_files} | New spit events: {ins_events}")
                if skipped:
                    st.dataframe(pd.DataFrame(skipped, columns=["File", "Reason"]), use_container_width=True)

# ---- Filters
st.subheader("🔎 Filters (combine as needed)")

today = date.today()
default_start = datetime.combine(today, time(0, 0, 0))
default_end = datetime.now().replace(microsecond=0)

c1, c2, c3, c4 = st.columns([1, 1, 1, 1])
with c1:
    start_date = st.date_input("Start date", value=default_start.date(), key="start_date")
    start_time = st.time_input("Start time", value=default_start.time(), key="start_time")
with c2:
    end_date = st.date_input("End date", value=default_end.date(), key="end_date")
    end_time = st.time_input("End time", value=default_end.time(), key="end_time")
with c3:
    board_value = st.number_input("Board value (for % loss)", min_value=0.0, value=0.0, step=1.0, key="board_value")
with c4:
    st.write("")
    st.write("")
    run_query = st.button("Run Query", type="primary")

dt_start = datetime.combine(start_date, start_time)
dt_end = datetime.combine(end_date, end_time)

# Filter option lists
boards_all = [r[0] for r in conn.execute(
    "SELECT DISTINCT board_name FROM logs WHERE board_name IS NOT NULL AND board_name <> '' ORDER BY board_name"
).fetchall()]
mos_all = [r[0] for r in conn.execute(
    "SELECT DISTINCT mo FROM logs WHERE mo IS NOT NULL AND mo <> '' ORDER BY mo"
).fetchall()]
machines_all = [r[0] for r in conn.execute(
    "SELECT DISTINCT machine FROM logs WHERE machine IS NOT NULL AND machine <> '' ORDER BY machine"
).fetchall()]
components_all = [r[0] for r in conn.execute(
    "SELECT DISTINCT component FROM events WHERE component IS NOT NULL AND component <> '' ORDER BY component"
).fetchall()]

# BOM selector
boms_df = list_boms(conn)
bom_labels, bom_id_by_label = [], {}
if not boms_df.empty:
    for _, r in boms_df.iterrows():
        label = f'{int(r["bom_id"])} | {r["bom_name"]} | {r["uploaded_at"]}'
        bom_labels.append(label)
        bom_id_by_label[label] = int(r["bom_id"])

f0, f1, f2, f3, f4 = st.columns([1.2, 1, 1, 1, 1])
with f0:
    selected_boms_labels = st.multiselect(
        "Master BOM version(s) to use for analysis (blank = latest per component)",
        options=bom_labels,
        default=[],
        key="selected_boms"
    )
with f1:
    boards_sel = st.multiselect("Board Name", boards_all, default=[], key="boards_sel")
with f2:
    mos_sel = st.multiselect("MO", mos_all, default=[], key="mos_sel")
with f3:
    machines_sel = st.multiselect("Machine", machines_all, default=[], key="machines_sel")
with f4:
    components_sel = st.multiselect("Component (optional)", components_all, default=[], key="components_sel")

selected_bom_ids = [bom_id_by_label[x] for x in selected_boms_labels] if selected_boms_labels else []

# Persist results
if "has_results" not in st.session_state:
    st.session_state.has_results = False

if run_query:
    bom_lookup = get_bom_lookup(conn, selected_bom_ids if selected_bom_ids else None)

    events_df = query_events(conn, dt_start, dt_end, boards_sel, mos_sel, machines_sel, components_sel, bom_lookup=bom_lookup)

    total_boards_est = estimate_total_boards(conn, dt_start, dt_end, boards_sel, mos_sel, machines_sel)

    boards_in_results = sorted([b for b in events_df["Board"].dropna().astype(str).unique()])
    boards_run_by_board = estimate_boards_by_board(conn, dt_start, dt_end, boards_sel, mos_sel, machines_sel, boards_limit=boards_in_results)

    m_breakdown = machine_log_breakdown(conn, dt_start, dt_end, boards_sel, mos_sel, machines_sel)

    summary_df = make_summary(events_df)
    repeated_df = make_repeated_locations(events_df)
    missing_df = make_missing_costs(events_df)
    board_loss_df = make_board_loss(events_df, board_value)
    board_loss_components_df = make_board_loss_components(events_df, boards_run_by_board, board_value)

    pareto_df = summary_df[["Component", "TotalCost"]].copy()
    pareto_df = pareto_df.sort_values("TotalCost", ascending=False).head(30)

    total_cost = float(events_df["Cost"].sum()) if not events_df.empty else 0.0
    yield_df = pd.DataFrame([
        ["Estimated Boards Run", round(float(total_boards_est), 3)],
        ["Total Cost Loss", round(total_cost, 2)],
        ["Avg Cost Loss / Board", round((total_cost / total_boards_est), 2) if total_boards_est else 0.0],
        ["Board Value (input)", float(board_value)],
    ], columns=["Metric", "Value"])

    st.session_state.payload = {
        "events_df": events_df,
        "summary_df": summary_df,
        "pareto_df": pareto_df,
        "repeated_df": repeated_df,
        "missing_df": missing_df,
        "board_loss_df": board_loss_df,
        "board_loss_components_df": board_loss_components_df,
        "yield_df": yield_df,
        "machine_breakdown": m_breakdown,
        "total_boards_est": float(total_boards_est),
        "total_cost": float(total_cost),
        "board_value": float(board_value),
    }
    st.session_state.has_results = True

# View selector
view = st.selectbox(
    "Select View",
    [
        "Summary",
        "Spit Events",
        "Pareto (Cost)",
        "Repeated Locations",
        "Yield Loss",
        "Missing BOM Costs",
        "Board Loss %",
        "Board Loss Components",
        "Chatbot (AI)"
    ],
    index=0
)

if not st.session_state.has_results:
    st.info("Upload BOM + ingest logs (once), then set filters and click **Run Query**.")
    st.stop()

payload = st.session_state.payload
events_df = payload["events_df"]
summary_df = payload["summary_df"]
pareto_df = payload["pareto_df"]
repeated_df = payload["repeated_df"]
missing_df = payload["missing_df"]
board_loss_df = payload["board_loss_df"]
board_loss_components_df = payload["board_loss_components_df"]
yield_df = payload["yield_df"]
machine_breakdown = payload["machine_breakdown"]

# ---- Excel Export button
with st.expander("⬇️ Export to Excel"):
    report_bytes = build_excel_report(
        events_df=events_df,
        summary_df=summary_df,
        pareto_df=pareto_df,
        repeated_df=repeated_df,
        yield_df=yield_df,
        missing_df=missing_df,
        board_loss_df=board_loss_df,
        board_loss_components_df=board_loss_components_df
    )
    fname = f"smt_spit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        "Download Excel Report",
        data=report_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Views
if view == "Summary":
    st.dataframe(summary_df, use_container_width=True)

elif view == "Spit Events":
    st.caption("Includes Feeder (col H) and Slot (col I), plus RejectCode.")
    st.dataframe(events_df, use_container_width=True)

elif view == "Pareto (Cost)":
    if events_df.empty:
        st.info("No events in this selection.")
    else:
        st.bar_chart(summary_df.set_index("Component")["TotalCost"].head(30))

elif view == "Repeated Locations":
    st.dataframe(repeated_df, use_container_width=True)

elif view == "Yield Loss":
    st.dataframe(yield_df, use_container_width=True)
    st.subheader("Machine log breakdown (used for board estimation)")
    st.dataframe(machine_breakdown, use_container_width=True)

elif view == "Missing BOM Costs":
    st.dataframe(missing_df, use_container_width=True)

elif view == "Board Loss %":
    st.dataframe(board_loss_df, use_container_width=True)

elif view == "Board Loss Components":
    st.dataframe(board_loss_components_df, use_container_width=True)

elif view == "Chatbot (AI)":
    st.subheader("Chatbot (Option 2: grounded answers + engineering suggestions)")

    if OpenAI is None:
        st.error("To enable the chatbot, add `openai` to requirements.txt and redeploy.")
        st.stop()

    api_key = st.secrets.get("OPENAI_API_KEY", None)
    if not api_key:
        st.warning('Add your key to Streamlit Secrets like:  OPENAI_API_KEY="sk-..."')
        st.stop()

    advisor_mode = st.toggle("Advisor mode (include engineering judgement)", value=True)
    show_debug = st.toggle("Show tool calls (debug)", value=False)

    default_start_iso = dt_start.isoformat(sep=" ")
    default_end_iso = dt_end.isoformat(sep=" ")

    # Chat state
    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = [
            {"role": "assistant", "content": "Ask me things like: 'What’s the top cost offender today and is it feeder-related?' or 'How many Board A did we run this week?'."}
        ]

    for m in st.session_state.chat_messages:
        with st.chat_message(m["role"]):
            st.markdown(m["content"])

    user_msg = st.chat_input("Ask anything about boards, rejects, feeder/slot, trends…")
    if user_msg:
        st.session_state.chat_messages.append({"role": "user", "content": user_msg})
        with st.chat_message("user"):
            st.markdown(user_msg)

        # Run chatbot
        reply = chatbot_reply(
            conn=conn,
            messages=st.session_state.chat_messages,
            default_start_iso=default_start_iso,
            default_end_iso=default_end_iso,
            advisor_mode=advisor_mode,
            selected_bom_ids=(selected_bom_ids if selected_bom_ids else None)
        )

        st.session_state.chat_messages.append({"role": "assistant", "content": reply.get("content", "")})

        with st.chat_message("assistant"):
            st.markdown(reply.get("content", ""))

            # Optional debug: show which tools were called
            if show_debug and reply.get("debug_calls"):
                with st.expander("Tool calls (debug)"):
                    st.json(reply["debug_calls"])
